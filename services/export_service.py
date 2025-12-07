"""
Export service for ReqCockpit
Handles exporting requirements and decisions to CSV and XLSX formats
"""
import logging
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime
import csv

from models.base import db_manager
from models.project import Project, Iteration, Supplier
from models.requirement import MasterRequirement, SupplierFeedback, CustREDecision
from config import EXPORT_FORMATS, MAX_EXCEL_ROWS, EXCEL_SHEET_NAME

logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for exporting project data to various formats
    
    Supports CSV and XLSX exports with flexible column selection
    and filtering options.
    """
    
    @staticmethod
    def export_to_csv(
        project_id: int,
        output_path: str,
        include_decisions: bool = True,
        include_all_iterations: bool = False,
        selected_suppliers: Optional[List[int]] = None
    ) -> Dict[str, Any]:
        """
        Export requirements and feedback to CSV format
        
        Args:
            project_id: ID of the project
            output_path: Path where CSV file will be written
            include_decisions: Whether to include CustRE decisions
            include_all_iterations: Include all iterations or just latest
            selected_suppliers: List of supplier IDs to include (None = all)
            
        Returns:
            Dictionary with export status and details
        """
        session = db_manager.get_session()
        if not session:
            return {
                'success': False,
                'message': 'No database connection',
                'rows_exported': 0
            }
        
        try:
            # Get project
            project = session.query(Project).filter(
                Project.id == project_id
            ).first()
            
            if not project:
                return {
                    'success': False,
                    'message': 'Project not found',
                    'rows_exported': 0
                }
            
            # Get all requirements
            requirements = session.query(MasterRequirement).filter(
                MasterRequirement.project_id == project_id
            ).all()
            
            if not requirements:
                return {
                    'success': False,
                    'message': 'No requirements found',
                    'rows_exported': 0
                }
            
            # Get suppliers
            suppliers = session.query(Supplier).filter(
                Supplier.project_id == project_id
            ).all()
            
            if selected_suppliers:
                suppliers = [s for s in suppliers if s.id in selected_suppliers]
            
            # Build header
            headers = ['ReqIF ID', 'Master Text']
            supplier_headers = [f'{s.name} (Status)' for s in suppliers]
            supplier_comment_headers = [f'{s.name} (Comment)' for s in suppliers]
            headers.extend(supplier_headers)
            headers.extend(supplier_comment_headers)
            
            if include_decisions:
                headers.extend(['Decision', 'Decision Note', 'Decision Date'])
            
            # Build rows
            rows = []
            for req in requirements:
                row = [req.reqif_id, req.master_text or '']
                
                # Add feedback for each supplier
                for supplier in suppliers:
                    feedback = session.query(SupplierFeedback).filter(
                        SupplierFeedback.requirement_id == req.id,
                        SupplierFeedback.supplier_id == supplier.id
                    ).order_by(
                        SupplierFeedback.created_at.desc()
                    ).first()
                    
                    if feedback:
                        row.append(feedback.normalized_status or '')
                    else:
                        row.append('')
                
                # Add comments
                for supplier in suppliers:
                    feedback = session.query(SupplierFeedback).filter(
                        SupplierFeedback.requirement_id == req.id,
                        SupplierFeedback.supplier_id == supplier.id
                    ).order_by(
                        SupplierFeedback.created_at.desc()
                    ).first()
                    
                    if feedback and feedback.comment:
                        row.append(feedback.comment)
                    else:
                        row.append('')
                
                # Add decision if requested
                if include_decisions:
                    decision = session.query(CustREDecision).filter(
                        CustREDecision.requirement_id == req.id
                    ).order_by(
                        CustREDecision.created_at.desc()
                    ).first()
                    
                    if decision:
                        row.append(decision.status)
                        row.append(decision.action_note or '')
                        row.append(
                            decision.created_at.isoformat() if decision.created_at else ''
                        )
                    else:
                        row.extend(['', '', ''])
                
                rows.append(row)
            
            # Write CSV
            try:
                output_file = Path(output_path)
                output_file.parent.mkdir(parents=True, exist_ok=True)
                
                with open(output_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)
                
                return {
                    'success': True,
                    'message': f'Exported {len(rows)} requirements to CSV',
                    'rows_exported': len(rows),
                    'file_path': str(output_file)
                }
            
            except IOError as e:
                logger.error(f"Failed to write CSV file: {e}")
                return {
                    'success': False,
                    'message': f'Failed to write file: {str(e)}',
                    'rows_exported': 0
                }
        
        finally:
            session.close()
    
    @staticmethod
    def export_to_xlsx(
        project_id: int,
        output_path: str,
        include_decisions: bool = True,
        selected_suppliers: Optional[List[int]] = None
    ) -> Dict[str, Any]:
        """
        Export requirements and feedback to XLSX format
        
        Args:
            project_id: ID of the project
            output_path: Path where XLSX file will be written
            include_decisions: Whether to include CustRE decisions
            selected_suppliers: List of supplier IDs to include (None = all)
            
        Returns:
            Dictionary with export status and details
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return {
                'success': False,
                'message': 'openpyxl library not installed',
                'rows_exported': 0
            }
        
        session = db_manager.get_session()
        if not session:
            return {
                'success': False,
                'message': 'No database connection',
                'rows_exported': 0
            }
        
        try:
            # Get project
            project = session.query(Project).filter(
                Project.id == project_id
            ).first()
            
            if not project:
                return {
                    'success': False,
                    'message': 'Project not found',
                    'rows_exported': 0
                }
            
            # Get all requirements
            requirements = session.query(MasterRequirement).filter(
                MasterRequirement.project_id == project_id
            ).all()
            
            if not requirements:
                return {
                    'success': False,
                    'message': 'No requirements found',
                    'rows_exported': 0
                }
            
            # Check row limit
            if len(requirements) > MAX_EXCEL_ROWS:
                return {
                    'success': False,
                    'message': f'Too many requirements ({len(requirements)}) for Excel export',
                    'rows_exported': 0
                }
            
            # Get suppliers
            suppliers = session.query(Supplier).filter(
                Supplier.project_id == project_id
            ).all()
            
            if selected_suppliers:
                suppliers = [s for s in suppliers if s.id in selected_suppliers]
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = EXCEL_SHEET_NAME
            
            # Build header
            headers = ['ReqIF ID', 'Master Text']
            supplier_headers = [f'{s.name} (Status)' for s in suppliers]
            supplier_comment_headers = [f'{s.name} (Comment)' for s in suppliers]
            headers.extend(supplier_headers)
            headers.extend(supplier_comment_headers)
            
            if include_decisions:
                headers.extend(['Decision', 'Decision Note', 'Decision Date'])
            
            # Write header
            ws.append(headers)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True)
            
            # Write data rows
            for req in requirements:
                row = [req.reqif_id, req.master_text or '']
                
                # Add feedback for each supplier
                for supplier in suppliers:
                    feedback = session.query(SupplierFeedback).filter(
                        SupplierFeedback.requirement_id == req.id,
                        SupplierFeedback.supplier_id == supplier.id
                    ).order_by(
                        SupplierFeedback.created_at.desc()
                    ).first()
                    
                    if feedback:
                        row.append(feedback.normalized_status or '')
                    else:
                        row.append('')
                
                # Add comments
                for supplier in suppliers:
                    feedback = session.query(SupplierFeedback).filter(
                        SupplierFeedback.requirement_id == req.id,
                        SupplierFeedback.supplier_id == supplier.id
                    ).order_by(
                        SupplierFeedback.created_at.desc()
                    ).first()
                    
                    if feedback and feedback.comment:
                        row.append(feedback.comment)
                    else:
                        row.append('')
                
                # Add decision if requested
                if include_decisions:
                    decision = session.query(CustREDecision).filter(
                        CustREDecision.requirement_id == req.id
                    ).order_by(
                        CustREDecision.created_at.desc()
                    ).first()
                    
                    if decision:
                        row.append(decision.status)
                        row.append(decision.action_note or '')
                        row.append(
                            decision.created_at.isoformat() if decision.created_at else ''
                        )
                    else:
                        row.extend(['', '', ''])
                
                ws.append(row)
            
            # Adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            try:
                output_file = Path(output_path)
                output_file.parent.mkdir(parents=True, exist_ok=True)
                wb.save(output_file)
                
                return {
                    'success': True,
                    'message': f'Exported {len(requirements)} requirements to XLSX',
                    'rows_exported': len(requirements),
                    'file_path': str(output_file)
                }
            
            except IOError as e:
                logger.error(f"Failed to write XLSX file: {e}")
                return {
                    'success': False,
                    'message': f'Failed to write file: {str(e)}',
                    'rows_exported': 0
                }
        
        finally:
            session.close()


# Global instance
export_service = ExportService()
